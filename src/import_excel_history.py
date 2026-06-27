"""
import_excel_history.py — import historical Excel snapshots into history.json.

Usage:
    python src/import_excel_history.py "path/to/file.xlsx" [--weekly]

Parses every block in the sheet that has the standard column layout:
    Owner | Total | Open | In Progress | Waiting For Approval | Overdue | Completed | Completion % | ...

Each block header row above the column names must contain the snapshot date
(either as a datetime cell or as YYYY-MM-DD text). Blocks without a parseable
date are skipped.

Merges into data/history.json — replaces existing entries for the same date,
keeps the file sorted by date.
"""
import os
import sys
import json
import warnings
from datetime import datetime, timezone
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

HISTORY_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "history.json")

REQUIRED_HEADERS = ["Owner", "Total", "Open", "In Progress",
                    "Waiting For Approval", "Overdue", "Completed", "Completion %"]


def _find_blocks(ws):
    """Find all (date, header_row, start_col) tuples in the sheet."""
    blocks = []
    for r in range(1, min(20, ws.max_row + 1)):  # headers usually in first 20 rows
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == "Owner":
                # Check next columns match expected headers
                ok = True
                for i, h in enumerate(REQUIRED_HEADERS):
                    cv = ws.cell(row=r, column=c + i).value
                    if cv != h:
                        ok = False
                        break
                if not ok:
                    continue
                # Find date in row above
                date = None
                for dr in range(max(1, r - 3), r):
                    dv = ws.cell(row=dr, column=c).value
                    if isinstance(dv, datetime):
                        date = dv.date()
                        break
                    if isinstance(dv, str):
                        try:
                            date = datetime.strptime(dv[:10], "%Y-%m-%d").date()
                            break
                        except ValueError:
                            continue
                if date:
                    blocks.append((date, r, c))
    return blocks


def _parse_block(ws, header_row, start_col):
    """Extract per-person stats from one block."""
    people = {}
    total_done = 0
    total_total = 0
    for r in range(header_row + 1, ws.max_row + 1):
        owner = ws.cell(row=r, column=start_col).value
        if owner is None or str(owner).strip() == "":
            break
        owner = str(owner).strip()
        if owner.upper() == "TOTAL":
            break
        total      = ws.cell(row=r, column=start_col + 1).value or 0
        open_      = ws.cell(row=r, column=start_col + 2).value or 0
        in_prog    = ws.cell(row=r, column=start_col + 3).value or 0
        wfa        = ws.cell(row=r, column=start_col + 4).value or 0
        overdue    = ws.cell(row=r, column=start_col + 5).value or 0
        completed  = ws.cell(row=r, column=start_col + 6).value or 0
        pct_raw    = ws.cell(row=r, column=start_col + 7).value

        # Pct may be 0-1 (Excel %) or 0-100
        if isinstance(pct_raw, (int, float)):
            pct = round(pct_raw * 100, 1) if pct_raw <= 1 else round(pct_raw, 1)
        elif total:
            pct = round(100.0 * completed / total, 1)
        else:
            pct = 0.0

        people[owner] = {
            "pct": pct,
            "done": int(completed),
            "total": int(total),
            "open": int(open_),
            "in_progress": int(in_prog),
            "waiting_for_approval": int(wfa),
            "overdue": int(overdue),
            "completed": int(completed),
            "statuses": {},
        }
        total_done += int(completed)
        total_total += int(total)

    team_pct = round(100.0 * total_done / total_total, 1) if total_total else 0.0
    return people, team_pct


def _load_history():
    if not os.path.exists(HISTORY_PATH):
        return []
    with open(HISTORY_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_history(history):
    history.sort(key=lambda s: s.get("date", ""))
    with open(HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)


def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    excel_path = sys.argv[1]
    mark_weekly = "--weekly" in sys.argv or True   # default to True for weekly Excel files

    print(f"Reading {excel_path} …")
    wb = load_workbook(excel_path, data_only=True)

    snapshots_added = 0
    history = _load_history()
    existing_dates = {s.get("date") for s in history}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        blocks = _find_blocks(ws)
        print(f"Sheet '{sheet_name}': found {len(blocks)} blocks")
        for date, hdr_row, start_col in blocks:
            people, team_pct = _parse_block(ws, hdr_row, start_col)
            iso_date = date.strftime("%Y-%m-%d")
            ts = datetime(date.year, date.month, date.day, 9, 0, 0,
                          tzinfo=timezone.utc).isoformat(timespec="seconds")
            snap = {
                "date": iso_date,
                "timestamp": ts,
                "team_total": team_pct,
                "is_weekly": mark_weekly,
                "people": people,
            }
            # Remove existing for this date, then append
            history = [s for s in history if s.get("date") != iso_date]
            history.append(snap)
            marker = "(replaced)" if iso_date in existing_dates else "(new)"
            print(f"  + {iso_date}  team={team_pct}%  people={len(people)}  {marker}")
            snapshots_added += 1

    _save_history(history)
    print(f"\nDone. Imported {snapshots_added} snapshot(s) into {HISTORY_PATH}")
    print(f"History now contains {len(history)} total snapshots.")


if __name__ == "__main__":
    main()
